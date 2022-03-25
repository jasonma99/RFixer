import os
import re
import numpy as np
import pandas as pd
import linecache
from natsort import natsorted
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def process_file(file_path, regex, counter, successful):
    """
    Compare the regex with positive and negative examples, return the number of matching positive examples,
    total number of positive exaples, the number of matching negative examples, total number of negative examples.

    :param file_path: the filepath of the test file.
    :param regex: the regex used to compare with examples
    :param counter: list of counters: [# pass solution regex check, # failed solution regex check, # successful==True]
    :param successful: boolean if the solution regex is successfully fixed.
    :return: nums=[pos_match, pos_total, neg_match, neg_total]
    """

    pos = False
    neg = False
    pos_match, pos_total = 0, 0
    neg_match, neg_total = 0, 0
    use_sol_regex = False

    with open(file_path, 'r') as infile:
        # use original regex if regex is None or successful is False
        if regex == "None" or successful == False:
            regex = infile.readline().rstrip("\n")
            pattern = re.compile(regex)
        else:
            try:
                # successful ==True solution regex is possible to fail to compile.
                pattern = re.compile(regex)
                use_sol_regex = True
            except Exception as e:
                # solution regex failed to compile, use original regex from test file.
                print(f"Fail to compile: {sheet_name}, {file_path}, {regex}")
                print(f"Exception message: {e}")
                sys.exit(1)
        for line in infile:
            line = line.rstrip("\n")
            if line == "+++":
                pos = True
                continue
            if line == "---":
                pos = False
                neg = True
                continue
            if pos:
                # if fullmatch does not return None
                if pattern.fullmatch(line):
                    pos_match += 1
                # elif use_sol_regex:
                #     print(f"Did not match: {sheet_name}, {file_path}, {regex}")
                pos_total += 1
            if neg:
                if not pattern.fullmatch(line):
                    neg_match += 1
                # elif use_sol_regex:
                #     print(f"Did not match: {sheet_name}, {file_path}, {regex}")
                neg_total += 1
    nums = np.asarray([pos_match, pos_total, neg_match, neg_total])

    # increase # of pass regex check only when all examples are matched.
    if use_sol_regex and pos_match == pos_total and neg_match == neg_total:
        counter[0] += 1
    # number of solution regex did not pass the check
    elif use_sol_regex and ( pos_match != pos_total or neg_match != neg_total ):
        counter[1] += 1

    return nums


def process_sheet(sheet, sheet_name, counter, tests_path, test_set):
    """
    Process a sheet of a dataset, save matching information to a text file.

    :param sheet: the excel sheet to be processed
    :param sheet_name: name of the excel sheet (experiment name)
    :param counter: counter list [# pass solution regex check, # failed solution regex check, # successful==True]
    :param tests_path: the path to "tests/" folder.
    :param test_set: the folder name of the testing dataset.
    :return total_nums: the 4 nums of the whole sheet.
    """

    total_nums = np.zeros(4)
    solution_regex_nums = np.zeros(4)
    for index, row in sheet.iterrows():
        file_name = row["filename"]
        regex = row["Solution regex"]
        successful = row["Successful"]
        file_path = os.path.join(tests_path, test_set, file_name)

        # process the test file with solution regex
        nums = process_file(file_path, regex, counter, successful)
        total_nums += nums
        if successful == True:
            counter[2] += 1
            solution_regex_nums += nums

    return total_nums, solution_regex_nums

def cal_f1(nums):
    """
    Calculate precision, recall, F1_score, F_score from given numbers.

    :param nums: the matching information: [pos_match, pos_total, neg_nonmatch, neg_total]
    :return precision, recall, F1_score, F_score
    """

    TP = nums[0]
    TN = nums[2]
    FP = nums[3] - nums[2]
    FN = nums[1] - nums[0]
    recall = TP / (TP + FN)
    precision = TP / (TP + FP)
    F1_score = 2 * precision * recall / ( precision + recall )
    F_score = (TP - FN + TN - FP) / (nums[1] + nums[3])

    return precision, recall, F1_score, F_score

def table1_definition():
    table = [["The total number of positive examples matched with solution regex from training dataset or initial regex from testing dataset."], 
             ["The total number of positive examples in the testing dataset."],
             ["The total number of negative examples did not match with solution regex from training dataset or initial regex from testing dataset."],
             ["The total number of negative examples in the testing dataset."],
             ["total pos match/total pos examples, the positive match ratio."],
             ["total neg mismatch/total neg examples, the negative mismatch ratio."],
             ["The number of solution regex from training dataset that passed check on testing data."],
             ["The number of solution regex from training dataset that failed check on testing data."],
             ["The number of solution regex returned by training."],
             ["Total number of regex in the dataset."],
             ["pass solution regex check/total Regex, the ratio of regex that passed check on testing data."],
             ["calculated using first 4 numbers. TP / (TP + FP)"],
             ["calculated using first 4 numbers. TP / (TP + FN)"],
             ["calculated using first 4 numbers. 2 * precision * recall / ( precision + recall )"],
             ["calculated using first 4 numbers. (TP - FN + TN - FP) / (total pos examples + total neg examples)"]]
    df = pd.DataFrame(table, columns=["Definition"])
    return df

def table2_definition():
    table = [["The total number of positive examples matched with solution regex from training dataset."], 
            ["The total number of positive examples for files in the testing dataset that has solution regex from training dataset."],
            ["The total number of negative examples did not match with solution regex from training dataset."],
            ["The total number of negative examples for files in the testing dataset that has solution regex from training dataset."],
            [""],
            [""],
            [""],
            [""],
            [""],
            ["pass solution regex check/RFixer Solution Regex, the ratio of solution regex that passed check on testing data."]]
    df = pd.DataFrame(table, columns=["Definition"])
    return df


def main():
    results_path = "results"
    tests_path = "tests"
    test_set = "dataset200"
    test_result = "results/dataset_test.xlsx"
    table_headers = ["total pos match", "total pos examples", "total neg mismatch", "total neg examples", "pos match ratio",
            "neg mismatch ratio", "pass solution regex check", "fail solution regex check", "RFixer Solution Regex",
            "total Regex in dataset", "pass rate (pass/total regex)", "precision", "recall", "F1 score", "F score"]
    table2_headers = ["total pos match", "total pos examples", "total neg mismatch", "total neg examples", "pos match ratio",
            "neg mismatch ratio", "pass solution regex check", "fail solution regex check", "RFixer Solution Regex",
            "pass rate (pass/solution regex)", "precision", "recall", "F1 score", "F score"]
    table = []
    table2 = []
    if os.path.isfile(test_result):
        writer = pd.ExcelWriter(test_result, mode="a", engine="openpyxl", if_sheet_exists="overlay")
    else:
        writer = pd.ExcelWriter(test_result, engine="openpyxl")

    ##########################################################################################
    # configuration
    excel_name = "dataset160_600"
    ##########################################################################################
    # read all sheets from the excel to a dictionary of dataframes: sheet_name=None
    df_dataset = pd.read_excel(f"results/{excel_name}.xlsx", sheet_name=None)
    sheet_names = list(df_dataset.keys())

    for sheet_name in sheet_names:
        sheet_dataset = df_dataset[sheet_name]
        # [# pass solution regex check, # failed solution regex check, # successful==True]
        counter = [0, 0, 0]
        # [pos_match, pos_total, neg_match, neg_total]
        sheet_nums, solution_regex_nums = process_sheet(sheet_dataset, sheet_name, counter, tests_path, test_set)

        # calculate matching ratio
        pos_ratio = sheet_nums[0] / sheet_nums[1]
        neg_ratio = sheet_nums[2] / sheet_nums[3]
        pos_ratio2 = solution_regex_nums[0] / solution_regex_nums[1]
        neg_ratio2 = solution_regex_nums[2] / solution_regex_nums[3]

        # calculate pass rate of regex
        total_regex = sheet_dataset.shape[0]
        pass_rate = counter[0] / total_regex
        pass_rate2 = counter[0] / counter[2]

        precision, recall, F1_score, F_score = cal_f1(sheet_nums)
        precision2, recall2, F1_score2, F_score2 = cal_f1(solution_regex_nums)

        # add new row of summary to summary table
        new_row = sheet_nums.tolist()
        new_row += [pos_ratio, neg_ratio, counter[0], counter[1], counter[2], total_regex, pass_rate, precision, recall, F1_score, F_score]
        table.append(new_row)

        new_row = solution_regex_nums.tolist()
        new_row += [pos_ratio2, neg_ratio2, counter[0], counter[1], counter[2], pass_rate2, precision2, recall2, F1_score2, F_score2]
        table2.append(new_row)
        print(f"not match {counter[1]} + match {counter[0]} = {counter[1] + counter[0]}, {counter[2]}")

    # Transpose and save summary table to excel sheet
    df = pd.DataFrame(table, columns = table_headers)
    df = df.T
    df.columns = sheet_names
    df.to_excel(writer, sheet_name=excel_name, index=True)
    # set column header width to the length of header
    worksheet = writer.sheets[excel_name]
    worksheet.column_dimensions["A"].width = len(table_headers[-5])
    for i in range(len(sheet_names)):
        worksheet.column_dimensions[get_column_letter(i+2)].width = len(sheet_names[i]) + 3
    # add table definition
    df = table1_definition()
    df.to_excel(writer, sheet_name=excel_name, index=False, startrow=0, startcol=len(sheet_names)+2)

    # write second table to same excel sheet
    df2 = pd.DataFrame(table2, columns=table2_headers)
    df2 = df2.T
    df2.columns = sheet_names
    df2.to_excel(writer, sheet_name=excel_name, index=True, startrow=df.shape[0] + 5, startcol=0)
    # add table definition
    df2 = table2_definition()
    df2.to_excel(writer, sheet_name=excel_name, index=False, startrow=df.shape[0] + 5, startcol=len(sheet_names)+2)

    writer.save()

if __name__ == "__main__":
    main()